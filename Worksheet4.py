from openpyxl import load_workbook
wb = load_workbook("/Users/dudleymac/Dropbox/Private/Book_Library.xlsx", read_only = True, data_only = True)
ws = wb.active
f1=open('./testfile', 'w+')

def pl(ptitle, pauthor, pread, pformat, pdate):
    print('{:115}'.format(ptitle), end=" ", file=f1)
    print('{:35}'.format(pauthor), end=" ", file=f1)
    print('{:8}'.format(pread), end=" ", file=f1)
    print('{:7}'.format(pformat), end=" ", file=f1)
    print('{:15}'.format(pdate), end=" ", file=f1)
    print(file=f1)

pl("Book Title", "Author", "Read", "Format", "Date Read")
print(file=f1)

kw = input("Search for: ")

for i in range (2, ws.max_row):
    if kw in ws.cell(row=i, column=1).value:
        if ws.cell(row=i, column=1).value == None:
            ptitle = " "
        else:
            ptitle = ws.cell(row=i, column=1).value

        if ws.cell(row=i, column=2).value == None:
            pauthor = " "
        else:
            pauthor = ws.cell(row=i, column=2).value

        if ws.cell(row=i, column=3).value == None:
            pread = " "
        else:
            pread = ws.cell(row=i, column=3).value

        if ws.cell(row=i, column=4).value == None:
            pformat = " "
        else:
            pformat = str(ws.cell(row=i, column=4).value)

        if ws.cell(row=i, column=5).value == None:
            pdate = " "
        else:
            pdate = str(ws.cell(row=i, column=5).value)

        pl(ptitle, pauthor, pread, pformat, pdate)

print(file=f1)
