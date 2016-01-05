from openpyxl import load_workbook
wb = load_workbook("/Users/dudleymac/Dropbox/Private/Book_Library.xlsx", read_only = True, data_only = True)
ws = wb.active
f1=open('./testfile', 'w+')

def pl(r):
    print('{:115}'.format(ws.cell(row=r, column=1).value), end=" ", file=f1)
    print('{:35}'.format(ws.cell(row=r, column=2).value), end=" ", file=f1)
    print('{:8}'.format(ws.cell(row=r, column=3).value), end=" ", file=f1)
    print('{:7}'.format(str(ws.cell(row=r, column=4).value)), end=" ", file=f1)
    print('{:15}'.format(str(ws.cell(row=r, column=5).value)), end=" ", file=f1)
    print(file=f1)

pl(1)
print(file=f1)

i=2
while ws.cell(row=i, column=1).value != None:
    pl(i)
    i += 1

print(file=f1)
