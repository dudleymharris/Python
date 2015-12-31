from openpyxl import load_workbook
wb = load_workbook("/Users/dudleymac/Dropbox/Private/Dudley's Book Library.xlsx", read_only = True, data_only = True)
ws = wb.active

print(ws['A1'].value+' '+ws['B1'].value+' '+ws['C1'].value+' '+ws['D1'].value+' '+ws['E1'].value)
